import streamlit as st

st.set_page_config(
    page_title="RAIO X OPERAÇÃO",
    page_icon="📌",
    layout="wide",
    initial_sidebar_state="expanded",
)

import pandas as pd
import numpy as np
import plotly.express as px
from datetime import datetime
from pathlib import Path
from io import BytesIO
from zoneinfo import ZoneInfo
import unicodedata
import re
from openpyxl import load_workbook

# =========================================================
# TÍTULO
# =========================================================
st.markdown(
    "<h2 style='text-align:center; margin-bottom:0;'>RAIO X</h2>"
    "<h3 style='text-align:center; margin-top:0;'>OPERAÇÃO</h3>",
    unsafe_allow_html=True
)

# =========================================================
# ARQUIVOS (pasta data/)
# Ajuste aqui APENAS os nomes dos arquivos se necessário
# =========================================================
DATA_DIR = Path(__file__).resolve().parents[1] / "data"

ARQ_ATIVOS = DATA_DIR / "Base colaboradores ativos.xlsx"              # BASE PRINCIPAL
ARQ_SKAP = DATA_DIR / "Skap.xlsx"                                     # pegar coluna NIVEIS
ARQ_PRONT = DATA_DIR / "Prontuario Condutor.xlsx"                     # Empregado + cor pintada em "Faixa"
ARQ_VALES = DATA_DIR / "Vales.xlsx"                                   # COLABORADOR + DATA
ARQ_RV = DATA_DIR / "RV.xlsx"                                         # NOME + DATA + TOTAL RV + % + RECARGA
ARQ_ABS = DATA_DIR / "Absenteismo.xlsx"                               # COLABORADOR + DATA + Tipo de Ausência
ARQ_ACID = DATA_DIR / "Ocorrencia de acidentes.xlsx"                  # COLABORADOR + Data da Ocorrência
ARQ_DTO = DATA_DIR / "Desvios de DTOs.xlsx"                           # COLABORADOR + Data
ARQ_IND = DATA_DIR / "Resultados indicadores operacionais.xlsx"       # COLABORADOR + Data + PDV + BEES + TML (MIN) + JL

# =========================================================
# ATUALIZAÇÃO + CACHE BUSTER + BOTÃO REFRESH
# =========================================================
def get_last_mtime():
    arquivos = [ARQ_ATIVOS, ARQ_SKAP, ARQ_PRONT, ARQ_VALES, ARQ_RV, ARQ_ABS, ARQ_ACID, ARQ_DTO, ARQ_IND]
    mtimes = [a.stat().st_mtime for a in arquivos if a.exists()]
    return max(mtimes) if mtimes else None

last_mtime = get_last_mtime()

try:
    if last_mtime is not None:
        dt = datetime.fromtimestamp(last_mtime, tz=ZoneInfo("America/Sao_Paulo"))
        st.caption(f"🕒 Última atualização dos dados: {dt.strftime('%d/%m/%Y %H:%M')}")
    else:
        st.caption("🕒 Última atualização: não disponível")
except Exception:
    st.caption("🕒 Última atualização: não disponível")

c_refresh, _ = st.columns([1, 5])
with c_refresh:
    if st.button("🔄 Atualizar dados agora", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

# =========================================================
# UTILS
# =========================================================
def normalizar_colunas(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [
        unicodedata.normalize("NFKD", str(c).strip()).encode("ascii", "ignore").decode("utf-8").upper()
        for c in df.columns
    ]
    return df

def normalizar_nome(s: str) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
    s = s.upper().strip()
    s = re.sub(r"[^\w\s]", " ", s)     # tira pontuação
    s = re.sub(r"\s+", " ", s).strip() # colapsa espaços
    return s

def tratar_data_segura(series: pd.Series) -> pd.Series:
    dt_txt = pd.to_datetime(series, errors="coerce", dayfirst=True)

    num = pd.to_numeric(series, errors="coerce").replace([np.inf, -np.inf], np.nan)
    mask = num.notna() & np.isfinite(num) & (num >= 20000) & (num <= 80000)

    dt_excel = pd.Series(pd.NaT, index=series.index)
    if mask.any():
        dt_excel.loc[mask] = pd.to_datetime(
            num.loc[mask].astype("int64"),
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )
    return dt_txt.fillna(dt_excel)

def to_month_key(dt: pd.Series) -> pd.Series:
    d = pd.to_datetime(dt, errors="coerce")
    return d.dt.to_period("M").astype(str)  # "YYYY-MM"

def month_label(pt_period: str) -> str:
    # "YYYY-MM" -> "MM/YYYY"
    try:
        y, m = pt_period.split("-")
        return f"{m}/{y}"
    except Exception:
        return pt_period

def centralizar_tabela(df: pd.DataFrame):
    return (
        df.style
        .set_properties(**{"text-align": "center"})
        .set_table_styles([{"selector": "th", "props": [("text-align", "center")]}])
    )

def preparar_excel_para_download(df: pd.DataFrame, sheet_name="Dados") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31])
    return output.getvalue()

def safe_float(x):
    return pd.to_numeric(x, errors="coerce")

def parse_percent(x):
    """
    Aceita:
    - 0.6117 -> 61.17%
    - "61,17%" / "61.17%" -> 61.17%
    - 61.17 -> 61.17%
    Retorna em porcentagem (0-100).
    """
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    if s == "":
        return np.nan
    s = s.replace("%", "").replace(",", ".")
    v = pd.to_numeric(s, errors="coerce")
    if pd.isna(v):
        return np.nan
    # se parece fração (0-1.5), converte para %
    if 0 <= v <= 1.5:
        v = v * 100
    return float(v)

def parse_time_mmss(x):
    """
    Retorna minutos como float.
    Aceita:
    - "00:30" / "0:30" (mm:ss)
    - "00:30:00" (hh:mm:ss)
    - número (já minutos)
    """
    if pd.isna(x):
        return np.nan
    s = str(x).strip()
    if s == "":
        return np.nan

    # número direto
    v = pd.to_numeric(s.replace(",", "."), errors="coerce")
    if not pd.isna(v) and re.fullmatch(r"[-+]?\d+(\.\d+)?", s.replace(",", ".")) is not None:
        return float(v)

    # tempo
    parts = s.split(":")
    parts = [p.strip() for p in parts if p.strip() != ""]
    if len(parts) == 2:
        mm, ss = parts
        mm = pd.to_numeric(mm, errors="coerce")
        ss = pd.to_numeric(ss, errors="coerce")
        if pd.isna(mm) or pd.isna(ss):
            return np.nan
        return float(mm) + float(ss)/60.0
    if len(parts) == 3:
        hh, mm, ss = parts
        hh = pd.to_numeric(hh, errors="coerce")
        mm = pd.to_numeric(mm, errors="coerce")
        ss = pd.to_numeric(ss, errors="coerce")
        if pd.isna(hh) or pd.isna(mm) or pd.isna(ss):
            return np.nan
        return float(hh)*60.0 + float(mm) + float(ss)/60.0

    return np.nan

# =========================================================
# FUNÇÕES (unificação)
# =========================================================
FUNCOES_MAP = {
    # Ajudante Distribuição/AS
    normalizar_nome("Ajudante Distribuição"): "Ajudante de Distribuição",
    normalizar_nome("Ajudante AS"): "Ajudante de Distribuição",

    # Motorista Distribuição
    normalizar_nome("Motorista Caminhão Distribuição"): "Motorista de Distribuição",
    normalizar_nome("Motorista de Distribuição AS"): "Motorista de Distribuição",

    # Motorista Van
    normalizar_nome("Motorista Entregador I"): "Motorista de Van",
    normalizar_nome("Motorista Entregador"): "Motorista de Van",
    normalizar_nome("Motorista Entregador II"): "Motorista de Van",

    # Ajudante Armazém
    normalizar_nome("Ajudante Armazém"): "Ajudante de armazem",
    normalizar_nome("Amarrador"): "Ajudante de armazem",

    # Operador
    normalizar_nome("Operador de Empilhadeira"): "Operador",
    normalizar_nome("Operador Conferente"): "Operador",
}

FUNCOES_PERMITIDAS = sorted(list(set(FUNCOES_MAP.values())))

def unificar_funcao(cargo: str) -> str:
    k = normalizar_nome(cargo)
    return FUNCOES_MAP.get(k, "")

# =========================================================
# METAS/PONTOS POR OPERAÇÃO
# (PDV em %; BEES em %; TML em mm:ss -> convertemos para minutos; JL em %)
# =========================================================
METAS = {
    "CD CASCAVEL": {
        "PDV": {"meta": 3.5, "pontos": 5, "tipo": "lte_pct"},
        "BEES": {"meta": 95.0, "pontos": 2, "tipo": "gte_pct"},
        "TML": {"meta": 0.5, "pontos": 2, "tipo": "lte_min"},  # 00:30 = 0.5 min
        "JL": {"meta": 80.0, "pontos": 3, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
    "CD DIADEMA": {
        "PDV": {"meta": 3.5, "pontos": 4, "tipo": "lte_pct"},
        "BEES": {"meta": 95.0, "pontos": 3, "tipo": "gte_pct"},
        "TML": {"meta": 0.5, "pontos": 3, "tipo": "lte_min"},
        "JL": {"meta": 80.0, "pontos": 1, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
    "CD FOZ DO IGUACU": {
        "PDV": {"meta": 3.01, "pontos": 3, "tipo": "lte_pct"},
        "BEES": {"meta": 95.0, "pontos": 2, "tipo": "gte_pct"},
        "TML": {"meta": 0.5, "pontos": 3, "tipo": "lte_min"},
        "JL": {"meta": 80.0, "pontos": 3, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
    "CD LONDRINA": {
        "PDV": {"meta": 3.5, "pontos": 3, "tipo": "lte_pct"},
        # BEES NÃO APLICÁVEL
        "TML": {"meta": 0.5, "pontos": 5, "tipo": "lte_min"},
        "JL": {"meta": 80.0, "pontos": 3, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
    "CD LITORAL": {
        "PDV": {"meta": 3.5, "pontos": 3, "tipo": "lte_pct"},
        "BEES": {"meta": 95.0, "pontos": 2, "tipo": "gte_pct"},
        "TML": {"meta": 0.5, "pontos": 3, "tipo": "lte_min"},
        "JL": {"meta": 80.0, "pontos": 3, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
    "CD SAO CRISTOVAO": {
        "PDV": {"meta": 3.5, "pontos": 3, "tipo": "lte_pct"},
        "BEES": {"meta": 95.0, "pontos": 4, "tipo": "gte_pct"},
        "TML": {"meta": 0.5, "pontos": 3, "tipo": "lte_min"},
        "JL": {"meta": 80.0, "pontos": 1, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
    "CD FRANCISCO BELTRAO": {
        "PDV": {"meta": 2.84, "pontos": 2, "tipo": "lte_pct"},
        "BEES": {"meta": 95.0, "pontos": 3, "tipo": "gte_pct"},
        "TML": {"meta": 0.5, "pontos": 3, "tipo": "lte_min"},
        "JL": {"meta": 89.0, "pontos": 3, "tipo": "gte_pct"},
        "ABS": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "VALES": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "ACIDENTE": {"meta": 0, "pontos": 5, "tipo": "eq0"},
        "DTO": {"meta": 0, "pontos": 4, "tipo": "eq0"},
    },
}

def norm_operacao(op: str) -> str:
    opn = normalizar_nome(op)
    opn = opn.replace("Ç", "C")
    return opn

# =========================================================
# PRONTUÁRIO: LER COR PINTADA (openpyxl)
# =========================================================
def excel_fill_to_hex(fill) -> str:
    """Tenta extrair RGB da cor de preenchimento."""
    try:
        if fill is None:
            return ""
        c = fill.fgColor
        if c is None:
            return ""
        if c.type == "rgb" and c.rgb:
            rgb = c.rgb
            # vem ARGB (ex: FF00FF00). pega só RGB
            if len(rgb) == 8:
                rgb = rgb[2:]
            return f"#{rgb}"
        return ""
    except Exception:
        return ""

def ler_prontuario_cores(path_xlsx: Path, col_nome_header="EMPREGADO", col_faixa_header="FAIXA") -> pd.DataFrame:
    """
    Retorna DF com: NOME_KEY, FAIXA_COR_HEX
    - Nome do colaborador: coluna "Empregado"
    - Cor: preenchimento (fill) da célula na coluna "Faixa"
    """
    wb = load_workbook(path_xlsx, data_only=True)
    ws = wb.active

    # acha colunas por header
    headers = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        h = normalizar_nome(str(cell.value))
        headers[h] = cell.column

    col_nome = headers.get(normalizar_nome(col_nome_header))
    col_faixa = headers.get(normalizar_nome(col_faixa_header))

    if not col_nome or not col_faixa:
        return pd.DataFrame(columns=["NOME_KEY", "PRONTUARIO_COR"])

    rows = []
    for r in range(2, ws.max_row + 1):
        nome = ws.cell(row=r, column=col_nome).value
        if nome is None or str(nome).strip() == "":
            continue
        nome_key = normalizar_nome(nome)

        cell_faixa = ws.cell(row=r, column=col_faixa)
        cor = excel_fill_to_hex(cell_faixa.fill)

        # também tenta pegar o texto (se existir)
        faixa_txt = "" if cell_faixa.value is None else str(cell_faixa.value).strip()
        rows.append({"NOME_KEY": nome_key, "PRONTUARIO_COR": cor, "FAIXA_TEXTO": faixa_txt})

    df = pd.DataFrame(rows)
    # se tiver duplicados, mantém o último
    if not df.empty:
        df = df.drop_duplicates(subset=["NOME_KEY"], keep="last")
    return df

# =========================================================
# LOAD (CACHE)
# =========================================================
@st.cache_data(show_spinner=True)
def carregar_bases(_cache_key: float | None):
    # base ativa é obrigatória
    if not ARQ_ATIVOS.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {ARQ_ATIVOS.name}")

    ativos = pd.read_excel(ARQ_ATIVOS)
    skap = pd.read_excel(ARQ_SKAP) if ARQ_SKAP.exists() else pd.DataFrame()
    vales = pd.read_excel(ARQ_VALES) if ARQ_VALES.exists() else pd.DataFrame()
    rv = pd.read_excel(ARQ_RV) if ARQ_RV.exists() else pd.DataFrame()
    abs_ = pd.read_excel(ARQ_ABS) if ARQ_ABS.exists() else pd.DataFrame()
    acid = pd.read_excel(ARQ_ACID) if ARQ_ACID.exists() else pd.DataFrame()
    dto = pd.read_excel(ARQ_DTO) if ARQ_DTO.exists() else pd.DataFrame()
    ind = pd.read_excel(ARQ_IND) if ARQ_IND.exists() else pd.DataFrame()

    pront = pd.DataFrame()
    if ARQ_PRONT.exists():
        try:
            pront = ler_prontuario_cores(ARQ_PRONT)
        except Exception:
            pront = pd.DataFrame()

    return ativos, skap, pront, vales, rv, abs_, acid, dto, ind

try:
    ativos, skap, pront, vales, rv, abs_, acid, dto, ind = carregar_bases(last_mtime)
except Exception as e:
    st.error(f"❌ Erro ao carregar bases: {e}")
    st.stop()

# =========================================================
# NORMALIZAÇÃO PRINCIPAL (ATIVOS)
# =========================================================
ativos = normalizar_colunas(ativos)

for c in ["COLABORADOR", "CARGO", "OPERACAO", "ATIVIDADE", "DATA ULT. ADM"]:
    if c not in ativos.columns:
        ativos[c] = ""

if "OPERAÇÃO" in ativos.columns and "OPERACAO" not in ativos.columns:
    ativos["OPERACAO"] = ativos["OPERAÇÃO"]

ativos["NOME_KEY"] = ativos["COLABORADOR"].astype(str).map(normalizar_nome)
ativos["OPER_KEY"] = ativos["OPERACAO"].astype(str).map(norm_operacao)
ativos["FUNCAO"] = ativos["CARGO"].astype(str).map(unificar_funcao)
ativos["DATA_ADM_DT"] = tratar_data_segura(ativos["DATA ULT. ADM"])
ativos["ADMISSAO"] = pd.to_datetime(ativos["DATA_ADM_DT"], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")

# Só funções permitidas
ativos = ativos[ativos["FUNCAO"].isin(FUNCOES_PERMITIDAS)].copy()

# =========================================================
# SKAP: puxar NÍVEIS (via nome normalizado)
# =========================================================
skap = normalizar_colunas(skap) if not skap.empty else skap
if not skap.empty:
    # tenta pegar "NIVEIS" + "COLABORADOR"
    if "COLABORADOR" not in skap.columns:
        skap["COLABORADOR"] = ""
    if "NIVEIS" not in skap.columns:
        skap["NIVEIS"] = ""
    skap["NOME_KEY"] = skap["COLABORADOR"].astype(str).map(normalizar_nome)
    skap_niveis = skap[["NOME_KEY", "NIVEIS"]].drop_duplicates(subset=["NOME_KEY"], keep="last")
else:
    skap_niveis = pd.DataFrame(columns=["NOME_KEY", "NIVEIS"])

# =========================================================
# PRONTUÁRIO: DF já vem (NOME_KEY, PRONTUARIO_COR, FAIXA_TEXTO)
# =========================================================
if pront is None or pront.empty:
    pront = pd.DataFrame(columns=["NOME_KEY", "PRONTUARIO_COR", "FAIXA_TEXTO"])

# =========================================================
# BASE MASTER (1 linha por colaborador) — para filtros e cards
# =========================================================
base_master = ativos.merge(skap_niveis, on="NOME_KEY", how="left").merge(pront, on="NOME_KEY", how="left")
for c in ["NIVEIS", "PRONTUARIO_COR", "FAIXA_TEXTO"]:
    if c not in base_master.columns:
        base_master[c] = ""

# =========================================================
# PREPARAR EVENTOS / OCORRÊNCIAS POR MÊS (tudo por NOME_KEY)
# =========================================================
def prep_evento(df: pd.DataFrame, col_nome: str, col_data: str, nome_out="NOME_KEY", data_out="DT"):
    if df is None or df.empty:
        return pd.DataFrame(columns=[nome_out, data_out, "MES"])
    df = normalizar_colunas(df)
    if col_nome.upper() not in df.columns:
        df[col_nome.upper()] = ""
    if col_data.upper() not in df.columns:
        df[col_data.upper()] = ""

    out = df[[col_nome.upper(), col_data.upper()]].copy()
    out[nome_out] = out[col_nome.upper()].astype(str).map(normalizar_nome)
    out[data_out] = tratar_data_segura(out[col_data.upper()])
    out = out.dropna(subset=[data_out])
    out["MES"] = to_month_key(out[data_out])
    return out[[nome_out, data_out, "MES"]]

# Vales: só contagem
vales_ev = prep_evento(vales, "COLABORADOR", "DATA")
# Acidentes: contagem
acid_ev = prep_evento(acid, "COLABORADOR", "DATA DA OCORRENCIA")
if acid_ev.empty:
    # alguns arquivos vêm com outro nome
    acid_ev = prep_evento(acid, "COLABORADOR", "DATA")
# DTO: contagem
dto_ev = prep_evento(dto, "COLABORADOR", "DATA")
# Abs: filtrar tipo ausência
abs_ev = pd.DataFrame(columns=["NOME_KEY", "DT", "MES", "TIPO"])
if abs_ is not None and not abs_.empty:
    abs_ = normalizar_colunas(abs_)
    for c in ["COLABORADOR", "DATA", "TIPO DE AUSENCIA"]:
        if c not in abs_.columns:
            abs_[c] = ""
    abs_["NOME_KEY"] = abs_["COLABORADOR"].astype(str).map(normalizar_nome)
    abs_["DT"] = tratar_data_segura(abs_["DATA"])
    abs_["TIPO"] = abs_["TIPO DE AUSENCIA"].astype(str).map(normalizar_nome)
    abs_ = abs_.dropna(subset=["DT"])
    abs_["MES"] = to_month_key(abs_["DT"])
    # somente JUSTIFICADA e NÃO JUSTIFICADA
    abs_ = abs_[abs_["TIPO"].isin([normalizar_nome("JUSTIFICADA"), normalizar_nome("NÃO JUSTIFICADA"), normalizar_nome("NAO JUSTIFICADA")])]
    abs_ev = abs_[["NOME_KEY", "DT", "MES", "TIPO"]].copy()

# RV: somar Total RV, média % e soma recarga (por mês)
rv_m = pd.DataFrame(columns=["NOME_KEY", "MES", "TOTAL_RV", "PCT_RV", "RECARGA"])
if rv is not None and not rv.empty:
    rv = normalizar_colunas(rv)
    # colunas esperadas
    if "NOME" not in rv.columns:
        rv["NOME"] = ""
    if "DATA" not in rv.columns:
        rv["DATA"] = ""
    # tenta achar nomes possíveis
    col_total = "TOTAL RV" if "TOTAL RV" in rv.columns else ("TOTAL_RV" if "TOTAL_RV" in rv.columns else "")
    col_pct = "%" if "%" in rv.columns else ("PCT" if "PCT" in rv.columns else ("PERCENTUAL" if "PERCENTUAL" in rv.columns else ""))
    col_rec = "RECARGA" if "RECARGA" in rv.columns else ""

    rv["NOME_KEY"] = rv["NOME"].astype(str).map(normalizar_nome)
    rv["DT"] = tratar_data_segura(rv["DATA"])
    rv = rv.dropna(subset=["DT"])
    rv["MES"] = to_month_key(rv["DT"])

    if col_total == "":
        rv["TOTAL_RV_NUM"] = np.nan
    else:
        rv["TOTAL_RV_NUM"] = safe_float(rv[col_total])

    if col_pct == "":
        rv["PCT_RV_NUM"] = np.nan
    else:
        rv["PCT_RV_NUM"] = rv[col_pct].apply(parse_percent)

    if col_rec == "":
        rv["RECARGA_NUM"] = np.nan
    else:
        rv["RECARGA_NUM"] = safe_float(rv[col_rec])

    rv_m = (
        rv.groupby(["NOME_KEY", "MES"], dropna=False)
        .agg(
            TOTAL_RV=("TOTAL_RV_NUM", "sum"),
            PCT_RV=("PCT_RV_NUM", "mean"),
            RECARGA=("RECARGA_NUM", "sum"),
        )
        .reset_index()
    )

# Indicadores operacionais: pegar resultado por mês
ind_m = pd.DataFrame(columns=["NOME_KEY", "MES", "PDV", "BEES", "TML_MIN", "JL"])
if ind is not None and not ind.empty:
    ind = normalizar_colunas(ind)
    for c in ["COLABORADOR", "DATA", "PDV", "BEES", "TML (MIN)", "JL"]:
        if c not in ind.columns:
            ind[c] = ""
    ind["NOME_KEY"] = ind["COLABORADOR"].astype(str).map(normalizar_nome)
    ind["DT"] = tratar_data_segura(ind["DATA"])
    ind = ind.dropna(subset=["DT"])
    ind["MES"] = to_month_key(ind["DT"])

    ind["PDV_NUM"] = ind["PDV"].apply(parse_percent)          # % (0-100)
    ind["BEES_NUM"] = ind["BEES"].apply(parse_percent)        # % (0-100)
    ind["TML_MIN"] = ind["TML (MIN)"].apply(parse_time_mmss)  # minutos
    ind["JL_NUM"] = ind["JL"].apply(parse_percent)            # % (0-100)

    # se tiver várias linhas no mês, pega média (padrão)
    ind_m = (
        ind.groupby(["NOME_KEY", "MES"], dropna=False)
        .agg(
            PDV=("PDV_NUM", "mean"),
            BEES=("BEES_NUM", "mean"),
            TML_MIN=("TML_MIN", "mean"),
            JL=("JL_NUM", "mean"),
        )
        .reset_index()
    )

# Contagens por mês:
vales_m = vales_ev.groupby(["NOME_KEY", "MES"]).size().reset_index(name="VALES")
acid_m = acid_ev.groupby(["NOME_KEY", "MES"]).size().reset_index(name="ACIDENTE")
dto_m = dto_ev.groupby(["NOME_KEY", "MES"]).size().reset_index(name="DTO")
abs_m = abs_ev.groupby(["NOME_KEY", "MES"]).size().reset_index(name="ABS")

# =========================================================
# TABELA BASE POR MÊS (colaborador x mês) — junta tudo
# =========================================================
# universo de meses = união de todos meses dos eventos + indicadores
meses_sets = []
for df_ in [vales_m, acid_m, dto_m, abs_m, rv_m, ind_m]:
    if df_ is not None and not df_.empty and "MES" in df_.columns:
        meses_sets.append(df_["MES"].dropna().unique().tolist())
meses_all = sorted(list(set([m for sub in meses_sets for m in sub])))

# se não houver mês em nenhum arquivo, cria mês atual
if not meses_all:
    meses_all = [pd.Timestamp.today().to_period("M").astype(str)]

# cross join: colaboradores (base_master) x meses_all
colab_univ = base_master[["NOME_KEY", "COLABORADOR", "OPERACAO", "OPER_KEY", "ATIVIDADE", "FUNCAO", "DATA_ADM_DT", "ADMISSAO", "NIVEIS", "PRONTUARIO_COR"]].copy()
colab_univ["OPER_KEY"] = colab_univ["OPER_KEY"].fillna(colab_univ["OPERACAO"].astype(str).map(norm_operacao))

grid = colab_univ.assign(key=1).merge(pd.DataFrame({"MES": meses_all, "key": 1}), on="key").drop(columns=["key"])

# joins
def left_join(grid_, df_, cols_keep):
    if df_ is None or df_.empty:
        for c in cols_keep:
            if c not in grid_.columns:
                grid_[c] = np.nan
        return grid_
    return grid_.merge(df_[["NOME_KEY", "MES"] + cols_keep], on=["NOME_KEY", "MES"], how="left")

grid = left_join(grid, ind_m, ["PDV", "BEES", "TML_MIN", "JL"])
grid = left_join(grid, abs_m, ["ABS"])
grid = left_join(grid, vales_m, ["VALES"])
grid = left_join(grid, acid_m, ["ACIDENTE"])
grid = left_join(grid, dto_m, ["DTO"])
grid = left_join(grid, rv_m, ["TOTAL_RV", "PCT_RV", "RECARGA"])

# preencher contagens nulas com 0
for c in ["ABS", "VALES", "ACIDENTE", "DTO"]:
    if c in grid.columns:
        grid[c] = pd.to_numeric(grid[c], errors="coerce").fillna(0).astype(int)

# =========================================================
# PONTUAÇÃO
# =========================================================
def calc_pontos_row(row) -> dict:
    op_key = norm_operacao(row.get("OPERACAO", ""))
    metas = METAS.get(op_key, None)

    res = {
        "PTS_PDV": 0, "PTS_BEES": 0, "PTS_TML": 0, "PTS_JL": 0,
        "PTS_ABS": 0, "PTS_VALES": 0, "PTS_ACIDENTE": 0, "PTS_DTO": 0,
        "TOTAL_PTS": 0
    }

    if metas is None:
        return res

    # PDV
    if "PDV" in metas:
        v = row.get("PDV", np.nan)
        if pd.notna(v) and v <= metas["PDV"]["meta"]:
            res["PTS_PDV"] = metas["PDV"]["pontos"]

    # BEES (pode não existir, exemplo Londrina)
    if "BEES" in metas:
        v = row.get("BEES", np.nan)
        if pd.notna(v) and v >= metas["BEES"]["meta"]:
            res["PTS_BEES"] = metas["BEES"]["pontos"]

    # TML
    if "TML" in metas:
        v = row.get("TML_MIN", np.nan)
        if pd.notna(v) and v <= metas["TML"]["meta"]:
            res["PTS_TML"] = metas["TML"]["pontos"]

    # JL
    if "JL" in metas:
        v = row.get("JL", np.nan)
        if pd.notna(v) and v >= metas["JL"]["meta"]:
            res["PTS_JL"] = metas["JL"]["pontos"]

    # ABS/VALES/ACIDENTE/DTO: meta 0
    if "ABS" in metas and int(row.get("ABS", 0)) == 0:
        res["PTS_ABS"] = metas["ABS"]["pontos"]
    if "VALES" in metas and int(row.get("VALES", 0)) == 0:
        res["PTS_VALES"] = metas["VALES"]["pontos"]
    if "ACIDENTE" in metas and int(row.get("ACIDENTE", 0)) == 0:
        res["PTS_ACIDENTE"] = metas["ACIDENTE"]["pontos"]
    if "DTO" in metas and int(row.get("DTO", 0)) == 0:
        res["PTS_DTO"] = metas["DTO"]["pontos"]

    res["TOTAL_PTS"] = sum([
        res["PTS_PDV"], res["PTS_BEES"], res["PTS_TML"], res["PTS_JL"],
        res["PTS_ABS"], res["PTS_VALES"], res["PTS_ACIDENTE"], res["PTS_DTO"]
    ])
    return res

pts_df = grid.apply(calc_pontos_row, axis=1, result_type="expand")
grid = pd.concat([grid, pts_df], axis=1)

def risco_to(total_pts: int) -> str:
    if total_pts > 20:
        return "NÃO"
    if total_pts >= 15:
        return "ATENÇÃO"
    return "SIM"

grid["RISCO DE TO"] = grid["TOTAL_PTS"].apply(lambda x: risco_to(int(x) if pd.notna(x) else 0))

# Label mês BR
grid["MÊS"] = grid["MES"].apply(month_label)

# =========================================================
# SIDEBAR — FILTROS
# =========================================================
st.sidebar.header("Filtros")

# Operação
ops_disp = sorted(base_master["OPERACAO"].dropna().astype(str).unique().tolist())
f_oper = st.sidebar.selectbox("Operação", ["Todos"] + ops_disp, index=0)

# Período (Mês/Ano baseado nos meses das ocorrências)
mes_opts = ["Todos"] + [month_label(m) for m in meses_all]
f_mes = st.sidebar.selectbox("Período", mes_opts, index=0)

# Função
f_func = st.sidebar.selectbox("Função", ["Todos"] + FUNCOES_PERMITIDAS, index=0)

# Atividade
ativs = sorted(base_master["ATIVIDADE"].dropna().astype(str).unique().tolist())
f_ativ = st.sidebar.selectbox("Atividade", ["Todos"] + ativs, index=0)

# Admissão (Início/Fim em PT-BR)
min_adm = pd.to_datetime(base_master["DATA_ADM_DT"], errors="coerce").min()
max_adm = pd.to_datetime(base_master["DATA_ADM_DT"], errors="coerce").max()
if pd.isna(min_adm) or pd.isna(max_adm):
    min_adm = pd.to_datetime("2024-01-01")
    max_adm = pd.to_datetime(datetime.today().date())

st.sidebar.subheader("Período de admissão")
col_ini, col_fim = st.sidebar.columns(2)
with col_ini:
    adm_ini = st.date_input(
        "Início",
        value=min_adm.date(),
        min_value=min_adm.date(),
        max_value=max_adm.date(),
        format="DD/MM/YYYY",
        key="raiox_adm_ini",
    )
with col_fim:
    adm_fim = st.date_input(
        "Fim",
        value=max_adm.date(),
        min_value=min_adm.date(),
        max_value=max_adm.date(),
        format="DD/MM/YYYY",
        key="raiox_adm_fim",
    )
if adm_ini > adm_fim:
    st.sidebar.warning("⚠️ Início maior que Fim. Ajustei automaticamente.")
    adm_ini, adm_fim = adm_fim, adm_ini

# Colaborador (depende dos filtros acima)
base_fil = base_master.copy()

if f_oper != "Todos":
    base_fil = base_fil[base_fil["OPERACAO"] == f_oper]
if f_func != "Todos":
    base_fil = base_fil[base_fil["FUNCAO"] == f_func]
if f_ativ != "Todos":
    base_fil = base_fil[base_fil["ATIVIDADE"].astype(str) == str(f_ativ)]

base_fil["DATA_ADM_DT"] = pd.to_datetime(base_fil["DATA_ADM_DT"], errors="coerce")
base_fil = base_fil[
    (base_fil["DATA_ADM_DT"].dt.date >= adm_ini) &
    (base_fil["DATA_ADM_DT"].dt.date <= adm_fim)
]

colabs_disp = sorted(base_fil["COLABORADOR"].dropna().astype(str).unique().tolist())
f_colab = st.sidebar.selectbox("Colaborador", ["Todos"] + colabs_disp, index=0)

# =========================================================
# APLICA FILTROS NA GRID
# =========================================================
df = grid.copy()

if f_oper != "Todos":
    df = df[df["OPERACAO"] == f_oper]
if f_func != "Todos":
    df = df[df["FUNCAO"] == f_func]
if f_ativ != "Todos":
    df = df[df["ATIVIDADE"].astype(str) == str(f_ativ)]

df["DATA_ADM_DT"] = pd.to_datetime(df["DATA_ADM_DT"], errors="coerce")
df = df[
    (df["DATA_ADM_DT"].dt.date >= adm_ini) &
    (df["DATA_ADM_DT"].dt.date <= adm_fim)
]

if f_mes != "Todos":
    # f_mes está em "MM/YYYY"
    # converte de volta pro "YYYY-MM" para filtrar
    mm, yy = f_mes.split("/")
    mes_key = f"{yy}-{mm}"
    df = df[df["MES"] == mes_key]

if f_colab != "Todos":
    nome_key = normalizar_nome(f_colab)
    df = df[df["NOME_KEY"] == nome_key]

# =========================================================
# BOX METAS/PONTUAÇÃO (lateral, abaixo filtros)
# =========================================================
st.sidebar.divider()
st.sidebar.subheader("Pontuação")
op_meta_key = norm_operacao(f_oper) if f_oper != "Todos" else None
metas_op = METAS.get(op_meta_key, None) if op_meta_key else None

if metas_op is None:
    st.sidebar.caption("Selecione uma operação com metas configuradas para ver a pontuação.")
else:
    # mostra metas e pontos conforme operação
    def linha_meta(nome, meta, pts):
        st.sidebar.markdown(f"- **{nome}**: meta **{meta}** · **{pts} pts**")

    if "PDV" in metas_op:
        linha_meta("PDV", f"≤ {metas_op['PDV']['meta']}%", metas_op["PDV"]["pontos"])
    if "BEES" in metas_op:
        linha_meta("BEES", f"≥ {metas_op['BEES']['meta']}%", metas_op["BEES"]["pontos"])
    if "TML" in metas_op:
        linha_meta("TML", "≤ 00:30", metas_op["TML"]["pontos"])
    if "JL" in metas_op:
        linha_meta("JL", f"≥ {metas_op['JL']['meta']}%", metas_op["JL"]["pontos"])
    linha_meta("Absenteísmo", "0", metas_op["ABS"]["pontos"])
    linha_meta("Vales", "0", metas_op["VALES"]["pontos"])
    linha_meta("Acidente", "0", metas_op["ACIDENTE"]["pontos"])
    linha_meta("Desvio DTO", "0", metas_op["DTO"]["pontos"])

    total_pts = sum(v["pontos"] for v in metas_op.values())
    st.sidebar.markdown(f"**Total possível:** {total_pts} pts")

# =========================================================
# TOPO — cards gerais / do colaborador
# =========================================================
# pega "linha referência" do colaborador (se selecionou) ou 1ª do filtro
ref = None
if not df.empty:
    # se tiver vários meses, usa o mais recente para mostrar dados fixos
    ref = df.sort_values("MES", ascending=False).iloc[0]

top1, top2, top3, top4 = st.columns([2.2, 1.2, 1.2, 1.2])

with top1:
    titulo_func = (ref["FUNCAO"] if ref is not None else "")
    st.markdown(f"<h4 style='margin:0;'> {titulo_func} </h4>", unsafe_allow_html=True)

with top2:
    if ref is None or pd.isna(ref.get("DATA_ADM_DT", None)):
        st.metric("Tempo de casa (meses)", "-")
    else:
        meses = int((pd.Timestamp.today().normalize() - pd.to_datetime(ref["DATA_ADM_DT"]).normalize()).days / 30.44)
        st.metric("Tempo de casa (meses)", f"{meses}")

with top3:
    st.metric("Admissão", (ref["ADMISSAO"] if ref is not None else "-"))

with top4:
    # base é ativos -> demissão normalmente não existe
    st.metric("Demissão", "-")

# =========================================================
# CARDS ESPECÍFICOS (Nível SKAP / Prontuário / etc.)
# =========================================================
cA, cB, cC, cD = st.columns(4)

with cA:
    st.metric("SKAP - Níveis", (ref["NIVEIS"] if ref is not None and str(ref.get("NIVEIS","")).strip() != "" else "-"))

with cB:
    cor_hex = (ref["PRONTUARIO_COR"] if ref is not None else "")
    if cor_hex:
        st.markdown(
            f"""
            <div style="padding:12px;border-radius:12px;background:#2b2f38;">
              <div style="font-weight:700;margin-bottom:6px;">Prontuário</div>
              <div style="display:flex;align-items:center;gap:10px;">
                <div style="width:18px;height:18px;border-radius:6px;background:{cor_hex};border:1px solid #555;"></div>
                <div style="opacity:0.95;">{cor_hex}</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True
        )
    else:
        st.metric("Prontuário", "-")

with cC:
    # placeholder (se você tiver colunas no ATIVOS, pode aparecer automático)
    # ex: "CICLO DE GENTE 2024" / "CICLO DE GENTE 2025"
    col = "CICLO DE GENTE 2025"
    if col in base_master.columns and ref is not None:
        st.metric("Ciclo de gente 2025", str(ref.get(col, "-")) or "-")
    else:
        st.metric("Ciclo de gente 2025", "-")

with cD:
    # média de pontuação no filtro atual (ou do colaborador se selecionado)
    if df.empty:
        st.metric("Pontuação (média)", "-")
    else:
        st.metric("Pontuação (média)", f"{df['TOTAL_PTS'].mean():.1f}")

st.divider()

# =========================================================
# GRÁFICO: pontuação total por mês (unidade ou colaborador)
# =========================================================
st.subheader("📊 Pontuação x mês")

graf = df.copy()
if graf.empty:
    st.info("Sem dados para o gráfico com os filtros atuais.")
else:
    g = (
        graf.groupby(["MES"], dropna=False)["TOTAL_PTS"]
        .mean()
        .reset_index()
        .sort_values("MES")
    )
    g["MÊS"] = g["MES"].apply(month_label)

    fig = px.bar(g, x="MÊS", y="TOTAL_PTS", text=g["TOTAL_PTS"].map(lambda x: f"{x:.0f}"))
    fig.update_layout(yaxis_title="Pontuação", xaxis_title="")
    st.plotly_chart(fig, use_container_width=True)

st.divider()

# =========================================================
# RANKING TOP 10
# =========================================================
st.subheader("🏆 TOP 10 melhores colaboradores (por pontuação média no filtro)")

if df.empty:
    st.info("Sem dados para ranking com os filtros atuais.")
else:
    rank = (
        df.groupby(["COLABORADOR"], dropna=False)
        .agg(
            OPERACAO=("OPERACAO", "first"),
            FUNCAO=("FUNCAO", "first"),
            MEDIA_PTS=("TOTAL_PTS", "mean"),
        )
        .reset_index()
        .sort_values("MEDIA_PTS", ascending=False)
        .head(10)
    )
    st.dataframe(centralizar_tabela(rank), use_container_width=True)

st.divider()

# =========================================================
# TABELA PRINCIPAL (modelo do exemplo)
# Mês / Nome / RISCO TO / indicadores + PTS + TOTAL
# =========================================================
st.subheader("📋 Tabela de Pontuação (por colaborador e mês)")

# colunas de resultado
def fmt_pct(v):
    return "" if pd.isna(v) else f"{v:.2f}%"

def fmt_min(v):
    if pd.isna(v):
        return ""
    # volta para mm:ss
    mm = int(v)
    ss = int(round((v - mm) * 60))
    if ss == 60:
        mm += 1
        ss = 0
    return f"{mm:02d}:{ss:02d}"

t = df.copy()

# BEES some no Londrina (mas pode aparecer vazio)
# cria colunas display
t["JORNADA"] = t.get("JL", np.nan).apply(fmt_pct)  # na sua imagem "Jornada" era um indicador; aqui usei JL como “Jornada”
t["PDV_TXT"] = t["PDV"].apply(fmt_pct)
t["BEES_TXT"] = t["BEES"].apply(fmt_pct)
t["TML_TXT"] = t["TML_MIN"].apply(fmt_min)
t["ABS_TXT"] = t["ABS"].astype(int).astype(str)
t["VALES_TXT"] = t["VALES"].astype(int).astype(str)
t["ACID_TXT"] = t["ACIDENTE"].astype(int).astype(str)
t["DTO_TXT"] = t["DTO"].astype(int).astype(str)

# pontos (para tabela)
def color_pts(v):
    try:
        v = int(v)
    except Exception:
        return ""
    return "color: #22c55e; font-weight:800;" if v > 0 else "color: #ef4444; font-weight:800;"

def color_risco(v):
    v = str(v)
    if v == "NÃO":
        return "color: #22c55e; font-weight:900;"
    if v == "ATENÇÃO":
        return "color: #facc15; font-weight:900;"
    if v == "SIM":
        return "color: #ef4444; font-weight:900;"
    return ""

# monta colunas finais
cols_out = [
    "MÊS",
    "COLABORADOR",
    "RISCO DE TO",
    "JORNADA", "PTS_JL",
    "PDV_TXT", "PTS_PDV",
    "BEES_TXT", "PTS_BEES",
    "TML_TXT", "PTS_TML",
    "ABS_TXT", "PTS_ABS",
    "VALES_TXT", "PTS_VALES",
    "ACID_TXT", "PTS_ACIDENTE",
    "DTO_TXT", "PTS_DTO",
    "TOTAL_PTS",
]

# remove BEES se a operação filtrada for Londrina (ou metas sem BEES)
if f_oper != "Todos":
    m = METAS.get(norm_operacao(f_oper), {})
    if "BEES" not in m:
        cols_out = [c for c in cols_out if c not in ["BEES_TXT", "PTS_BEES"]]

t_out = t[cols_out].copy()

# ordenação
t_out["_MESKEY"] = pd.to_datetime(t["MES"] + "-01", errors="coerce")
t_out = t_out.sort_values(["_MESKEY", "COLABORADOR"]).drop(columns=["_MESKEY"])

sty = centralizar_tabela(t_out)

# aplica cor em pontos
pts_cols = [c for c in t_out.columns if c.startswith("PTS_")]
if pts_cols:
    sty = sty.applymap(color_pts, subset=pts_cols)

# aplica cor no risco
if "RISCO DE TO" in t_out.columns:
    sty = sty.applymap(color_risco, subset=["RISCO DE TO"])

# total em destaque
sty = sty.applymap(lambda v: "font-weight:900;" if isinstance(v, (int, float, np.integer, np.floating)) else "", subset=["TOTAL_PTS"])

st.dataframe(sty, use_container_width=True, height=520)

# Export
excel = preparar_excel_para_download(t_out, sheet_name="RAIO_X")
st.download_button(
    "⬇️ Baixar Excel (Tabela RAIO X)",
    data=excel,
    file_name="raio_x_operacao.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True
)

# =========================================================
# NOTAS IMPORTANTES (bem curtas)
# =========================================================
with st.expander("ℹ️ Notas rápidas"):
    st.write(
        "- O match entre planilhas é feito por **nome normalizado** (sem acento, sem pontuação, espaços ajustados).\n"
        "- Prontuário: a cor é lida do **preenchimento da célula** na coluna **Faixa**.\n"
        "- Indicadores por mês: quando há mais de um registro no mês, o app usa **média** (PDV/BEES/TML/JL)."
    )
